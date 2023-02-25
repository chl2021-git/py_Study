import os
import openpyxl

folder_path = "path/to/folder"

for file_name in os.listdir(folder_path):
    if file_name.endswith(".xlsx"):
        file_path = os.path.join(folder_path, file_name)
        workbook = openpyxl.load_workbook(file_path)

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and "abc" in str(cell.value):
                        cell.value = str(cell.value).replace("abc", "def")

        workbook.save(file_path)
